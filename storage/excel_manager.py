"""
Excel Manager Module
Handles reading and writing hackathon data to Excel files.
"""

import os
import logging
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

class ExcelManager:
    def __init__(self, excel_file_path):
        self.excel_file = Path(excel_file_path)
        self.logger = logging.getLogger(__name__)
        self.headers = [
            'Name', 'Platform', 'Link', 'Start Date', 
            'Tags', 'Scraped At', 'Status'
        ]
        self.ensure_excel_file()
        
    def ensure_excel_file(self):
        """Create Excel file if it doesn't exist"""
        if not self.excel_file.exists():
            self.create_new_excel_file()
        else:
            self.validate_excel_structure()
            
    def create_new_excel_file(self):
        """Create a new Excel file with proper headers"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Hackathons"
            
            # Add headers
            for col, header in enumerate(self.headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
                
            # Set column widths
            column_widths = [40, 15, 50, 15, 30, 20, 15]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
                
            wb.save(self.excel_file)
            self.logger.info(f"Created new Excel file: {self.excel_file}")
            
        except Exception as e:
            self.logger.error(f"Error creating Excel file: {e}")
            raise
            
    def validate_excel_structure(self):
        """Validate that the Excel file has the correct structure"""
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            # Check if headers exist
            if ws.max_row == 0:
                self.create_new_excel_file()
                return
                
            # Validate headers
            existing_headers = [ws.cell(row=1, column=col).value for col in range(1, len(self.headers) + 1)]
            if existing_headers != self.headers:
                self.logger.warning("Excel file headers don't match expected format")
                
        except Exception as e:
            self.logger.error(f"Error validating Excel structure: {e}")
            self.create_new_excel_file()
            
    def get_existing_hackathons(self):
        """Get list of existing hackathons from Excel file"""
        hackathons = []
        try:
            if not self.excel_file.exists():
                return hackathons
                
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            # Skip header row
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # If name is not empty
                    hackathon = {
                        'name': row[0],
                        'platform': row[1],
                        'link': row[2],
                        'start_date': row[3],
                        'tags': row[4],
                        'scraped_at': row[5],
                        'status': row[6] if len(row) > 6 else 'New'
                    }
                    hackathons.append(hackathon)
                    
        except Exception as e:
            self.logger.error(f"Error reading existing hackathons: {e}")
            
        return hackathons
        
    def save_hackathons(self, hackathons):
        """Save new hackathons to Excel file"""
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            # Find the next empty row
            next_row = ws.max_row + 1
            
            for hackathon in hackathons:
                ws.cell(row=next_row, column=1, value=hackathon.get('name', ''))
                ws.cell(row=next_row, column=2, value=hackathon.get('platform', ''))
                ws.cell(row=next_row, column=3, value=hackathon.get('link', ''))
                ws.cell(row=next_row, column=4, value=hackathon.get('start_date', ''))
                ws.cell(row=next_row, column=5, value=hackathon.get('tags', ''))
                ws.cell(row=next_row, column=6, value=hackathon.get('scraped_at', ''))
                ws.cell(row=next_row, column=7, value='New')
                
                next_row += 1
                
            wb.save(self.excel_file)
            self.logger.info(f"Saved {len(hackathons)} hackathons to Excel file")
            
        except Exception as e:
            self.logger.error(f"Error saving hackathons to Excel: {e}")
            raise
            
    def update_hackathon_status(self, hackathon_name, status):
        """Update the status of a specific hackathon"""
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=1).value == hackathon_name:
                    ws.cell(row=row, column=7, value=status)
                    break
                    
            wb.save(self.excel_file)
            self.logger.info(f"Updated status for '{hackathon_name}' to '{status}'")
            
        except Exception as e:
            self.logger.error(f"Error updating hackathon status: {e}")
            
    def get_hackathon_stats(self):
        """Get statistics about stored hackathons"""
        try:
            hackathons = self.get_existing_hackathons()
            
            stats = {
                'total': len(hackathons),
                'platforms': {},
                'recent': 0
            }
            
            # Count by platform
            for hackathon in hackathons:
                platform = hackathon.get('platform', 'Unknown')
                stats['platforms'][platform] = stats['platforms'].get(platform, 0) + 1
                
            # Count recent (last 7 days)
            recent_date = datetime.now().strftime('%Y-%m-%d')
            for hackathon in hackathons:
                scraped_at = hackathon.get('scraped_at', '')
                if recent_date in scraped_at:
                    stats['recent'] += 1
                    
            return stats
            
        except Exception as e:
            self.logger.error(f"Error getting hackathon stats: {e}")
            return {'total': 0, 'platforms': {}, 'recent': 0}
